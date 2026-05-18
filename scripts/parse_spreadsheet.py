import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize_cell(value):
    if value is None:
      return ""
    return str(value).strip()


def trim_row(row):
    values = [normalize_cell(cell) for cell in row]
    while values and values[-1] == "":
        values.pop()
    return values


def main():
    if len(sys.argv) < 2:
        raise SystemExit("usage: parse_spreadsheet.py <xlsx-path>")

    workbook = load_workbook(filename=Path(sys.argv[1]), read_only=True, data_only=True)
    worksheet = workbook.active

    rows = []
    for row in worksheet.iter_rows(values_only=True):
        trimmed = trim_row(row)
        if any(cell != "" for cell in trimmed):
            rows.append(trimmed)

    print(
        json.dumps(
            {
                "sheetName": worksheet.title,
                "rows": rows,
            }
        )
    )


if __name__ == "__main__":
    main()
